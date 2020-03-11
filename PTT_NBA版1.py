# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import openpyxl
from requests import get
from openpyxl import Workbook
import urllib.request as req
def geturl(url):
    from bs4 import BeautifulSoup
    request=req.Request(url, headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
        })       
    with req.urlopen(request) as response:
        data=response.read().decode("utf-8")
    soup = BeautifulSoup(data, 'html.parser')  
    nextLink=soup.find("a",string="‹ 上頁")
    return nextLink["href"]
  
def getdata(url):
    from bs4 import BeautifulSoup
    #建立一個request物件 附加request header的資訊
    request=req.Request(url, headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
        })
    with req.urlopen(request) as response:
        data=response.read().decode("utf-8")
    soup = BeautifulSoup(data, 'html.parser')    
    #尋找所有class=title的div標籤
    titles=soup.find_all("div",class_="title")
    titles_webs=soup.find_all("div",class_="title")
    #titles_web= [b.a["href"] for b in titles_webs] 
     # 創建一個空白活頁簿物件
    date=soup.find_all('div','date')
    date_number= [e.get_text() for e in date]
    push=soup.find_all('div','nrec')
    push_number= [a.get_text() for a in push]
    author=soup.find_all('div','author')
    author_number= [c.get_text() for c in author]
    i=0
    for title in titles:
        if author_number[i] =='-':    
            print("文章已刪除",0,0,0)
            ws.append(["文章已刪除",0,0,0])
            i=i+1
        else:
            print(title.a.string,date_number[i],author_number[i],push_number[i],gettitledata("https://www.ptt.cc"+titles_webs[i].a["href"]))
            ws.append([title.a.string,date_number[i],author_number[i],push_number[i],gettitledata("https://www.ptt.cc"+titles_webs[i].a["href"])])
            i=i+1
    #抓上一頁連結  
def gettitledata(url):
    from bs4 import BeautifulSoup 
    #建立一個request物件 附加request header的資訊
    request=req.Request(url, headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
                                      })
    with req.urlopen(request) as response:
        data=response.read().decode("utf-8")
    soup = BeautifulSoup(data, 'html.parser')
#尋找所有class=title的div標籤
    
    request=req.Request(url, headers={"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.122 Safari/537.36"
                                                 })
    with req.urlopen(request) as response:
        data=response.read().decode("utf-8")
    soup = BeautifulSoup(data, 'html.parser')
    nopush=soup.find_all("span","f1 hl push-tag")
    nopushs= [a.get_text() for a in nopush]
    nopush_number=0
    i=0
    for a in nopushs:
        if nopushs[i]=='噓 ':
            nopush_number=nopush_number+1
        i=i+1    
    return nopush_number
    
      
wb = Workbook()
# 選取正在工作中的表單
ws = wb.active
ws['A1'] = '標題'     
ws['B1'] = '日期' 
ws['C1'] = '作者' 
ws['D1'] = '推文數'    
ws['E1'] = '噓數' 
url="https://www.ptt.cc/bbs/NBA/index.html"

for i in range(5):
    getdata(url)
    url="https://www.ptt.cc"+geturl(url)
    

# 儲存成 create_sample.xlsx 檔案
wb.save('PTT_NBA版.xlsx')
# 儲存成XLSX檔